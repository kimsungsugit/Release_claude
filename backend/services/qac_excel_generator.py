"""QAC Excel 리포트 생성기

QAC 리포트를 Excel 형식으로 생성합니다.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional

from backend.services.qac_parser import QACDataManager, MatrixItem, HISItem
from backend.services.vcast_excel_generator import XlsxManager, XlsCellStyle

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    BarChart = None
    PieChart = None
    Reference = None
    DataPoint = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    get_column_letter = None


def generate_qac_excel(qac_manager: QACDataManager, output_path: Path) -> bool:
    """QAC Excel 리포트 생성
    
    Args:
        qac_manager: 파싱된 QAC 데이터
        output_path: 출력 파일 경로
    
    Returns:
        성공 여부
    """
    if Workbook is None:
        raise ImportError("openpyxl is required")
    
    excel = XlsxManager()
    if not excel.create(output_path):
        return False
    
    col_offset = 1
    row_offset = 3
    title_col_count = 7
    
    excel.select_sheet(1, "QAC Report", True)

    # 스펙 카운트 리셋 (Excel 생성 시 다시 카운트)
    qac_manager.clear_spec_over_count()

    # 제목
    excel.write_data(1, 1, "QAC Report")
    excel.apply_style(1, 1, 1, title_col_count, XlsCellStyle.Title)
    excel.merge(1, 1, 1, title_col_count)
    
    # 헤더 설정
    matrix_list = QACDataManager.get_matrix_list()
    col_count = 2 + len(matrix_list) + 1  # Index, Function, Matrix items, File
    
    # 헤더 행 1
    current_row = row_offset
    excel.write_data(current_row, col_offset, "Index")
    excel.write_data(current_row, col_offset + 1, "Function")
    
    col_idx = col_offset + 2
    for matrix in matrix_list:
        title = HISItem.get_title(matrix, True)
        excel.write_data(current_row, col_idx, title)
        col_idx += 1
    
    excel.write_data(current_row, col_idx, "File")
    
    # 헤더 행 2
    current_row += 1
    excel.write_data(current_row, col_offset, "Index")
    excel.write_data(current_row, col_offset + 1, "Function")
    
    col_idx = col_offset + 2
    for matrix in matrix_list:
        title = HISItem.get_title(matrix, False)
        excel.write_data(current_row, col_idx, title)
        col_idx += 1
    
    excel.write_data(current_row, col_idx, "File")
    
    # 헤더 스타일 적용
    excel.apply_style(row_offset, col_offset, row_offset + 1, col_offset + col_count - 1, XlsCellStyle.Caption)
    
    # 헤더 병합: Index, Function, File은 2행 병합 (C# 원본과 동일)
    excel.merge(row_offset, col_offset, row_offset + 1, col_offset)          # Index
    excel.merge(row_offset, col_offset + 1, row_offset + 1, col_offset + 1)  # Function
    # File column: NOT merged, both rows have "File" (C# compatible)

    # 데이터 행
    current_row += 1
    row_num = 0  # C# 원본은 0부터 시작

    for his_item in qac_manager.list_result:
        col = col_offset
        excel.write_data(current_row, col, row_num)
        col += 1
        # C# 원본은 함수명/파일명 앞에 space 보존
        excel.write_data(current_row, col, his_item.function_name)
        col += 1
        
        # Matrix 값들
        for matrix in matrix_list:
            value = his_item.get_matrix_value(matrix)
            warning_level = qac_manager.check_warning_level(matrix, value)
            
            excel.write_data(current_row, col, value)
            
            # 경고 레벨에 따른 스타일 적용
            if warning_level == 1:
                excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgYellow)
            elif warning_level == 2:
                excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgOrange)
            elif warning_level == 3:
                excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgRed)
            
            qac_manager.update_spec_over_count(matrix, warning_level)
            col += 1
        
        excel.write_data(current_row, col, his_item.file_name)
        
        excel.set_wrap_text(current_row, col_offset, current_row, col_offset + col_count - 1, True)
        current_row += 1
        row_num += 1
    
    # Total 행들 (Level 0 = 경고 없음, Level 1,2,3 = 경고 레벨별)
    total_items = len(qac_manager.list_result)
    for warn_level in range(0, 4):  # Level 0, 1, 2, 3
        if warn_level == 0:
            excel.write_data(current_row, col_offset, "Total")
        excel.write_data(current_row, col_offset + 1, f"Level {warn_level}")
        
        col = col_offset + 2
        spec_string = ""
        
        for matrix in matrix_list:
            col_idx = qac_manager.get_column_index_of_matrix_item(matrix)
            if col_idx < 0:
                excel.write_data(current_row, col, "-")
                excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgLightGray)
            elif warn_level == 0:
                # Level 0: directly from spec_over_count[0] (C# compatible)
                if matrix in qac_manager.dic_spec_over_count:
                    spec = qac_manager.dic_spec_over_count[matrix]
                    count = spec.list_spec[0] if len(spec.list_spec) > 0 else 0
                    excel.write_data(current_row, col, str(count))
                else:
                    excel.write_data(current_row, col, str(total_items))
            else:
                if matrix in qac_manager.dic_spec_over_count:
                    spec = qac_manager.dic_spec_over_count[matrix]
                    if warn_level < len(spec.list_spec):
                        count = spec.list_spec[warn_level]
                        excel.write_data(current_row, col, str(count))

                        if warn_level == 1:
                            excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgYellow)
                        elif warn_level == 2:
                            excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgOrange)
                        elif warn_level == 3:
                            excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgRed)

                        spec_str = qac_manager.get_spec_string(matrix, warn_level)
                        if spec_str:
                            if spec_string:
                                spec_string += ", "
                            spec_string += spec_str
                    else:
                        excel.write_data(current_row, col, "-")
                        excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgLightGray)
                else:
                    excel.write_data(current_row, col, "-")
                    excel.apply_style(current_row, col, current_row, col, XlsCellStyle.BgLightGray)

            col += 1
        
        excel.write_data(current_row, col, spec_string)
        excel.apply_style(current_row, col_offset, current_row, col, XlsCellStyle.BgLightGray)
        current_row += 1
    
    # Total 행의 "Total" 셀 병합 (4개 행)
    total_start_row = current_row - 4
    excel.merge(total_start_row, col_offset, current_row - 1, col_offset)

    # 일반 데이터 스타일 적용
    excel.apply_style(row_offset + 2, col_offset, current_row - 1, col_offset + col_count - 1, XlsCellStyle.General)

    # 열 너비 설정 (C# 출력 기준: A=80, B=300, C=80, D/E/F=60, G=500)
    widths = [80, 300, 80, 60, 60, 60, 500]
    for col in range(col_offset, col_offset + col_count):
        if col - col_offset < len(widths):
            excel.set_column_width(col, widths[col - col_offset])
    
    # Summary 차트 시트 추가
    add_summary_charts(excel.workbook, qac_manager)

    return excel.close(True)


def add_summary_charts(wb, manager: QACDataManager) -> None:
    """QAC Summary 시트에 차트를 생성합니다.

    Args:
        wb: openpyxl Workbook 인스턴스
        manager: 파싱된 QAC 데이터 매니저
    """
    if BarChart is None or not manager.list_result:
        return

    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------------
    # 1. 타이틀 & 요약 통계
    # ------------------------------------------------------------------
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "QAC Summary"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "Total functions"
    ws["B3"] = len(manager.list_result)
    ws["A3"].font = Font(bold=True)

    # ------------------------------------------------------------------
    # 2. v(G) 복잡도 분포 데이터 테이블 (bar chart 원본)
    #    각 함수에 대해 warning level 0/1/2/3 중 하나를 매김
    # ------------------------------------------------------------------
    bar_data_start_row = 6
    ws.cell(row=bar_data_start_row, column=1, value="Function")
    ws.cell(row=bar_data_start_row, column=2, value="v(G)")
    ws.cell(row=bar_data_start_row, column=3, value="Warning Level")
    for c in range(1, 4):
        ws.cell(row=bar_data_start_row, column=c).font = Font(bold=True)

    data_row = bar_data_start_row + 1
    for his_item in manager.list_result:
        vg_value = his_item.get_matrix_value(MatrixItem.V_G)
        warn_level = manager.check_warning_level(MatrixItem.V_G, vg_value)

        ws.cell(row=data_row, column=1, value=his_item.function_name)

        # v(G) 값을 숫자로 기록 (차트에서 사용)
        try:
            ws.cell(row=data_row, column=2, value=int(vg_value))
        except (ValueError, TypeError):
            ws.cell(row=data_row, column=2, value=0)

        ws.cell(row=data_row, column=3, value=warn_level)
        data_row += 1

    bar_data_end_row = data_row - 1

    # --- Bar Chart: v(G) complexity per function ---
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "v(G) Complexity Distribution"
    bar_chart.y_axis.title = "Cyclomatic Complexity"
    bar_chart.x_axis.title = "Function"

    cats = Reference(ws, min_col=1, min_row=bar_data_start_row + 1, max_row=bar_data_end_row)
    vals = Reference(ws, min_col=2, min_row=bar_data_start_row, max_row=bar_data_end_row)
    bar_chart.add_data(vals, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.shape = 4
    bar_chart.width = 30
    bar_chart.height = 15

    # 색상 매핑: warning level -> fill colour
    level_colors = {0: "00B050", 1: "FFFF00", 2: "FFA500", 3: "FF0000"}
    series = bar_chart.series[0]
    for idx in range(bar_data_end_row - bar_data_start_row):
        level_val = ws.cell(row=bar_data_start_row + 1 + idx, column=3).value
        color = level_colors.get(level_val, "00B050")
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)

    ws.add_chart(bar_chart, "E3")

    # ------------------------------------------------------------------
    # 3. 준수율 파이 차트 데이터 테이블
    #    Pass (level 0) vs Yellow (1) vs Orange (2) vs Red (3)
    # ------------------------------------------------------------------
    pie_data_start_row = bar_data_end_row + 3
    ws.cell(row=pie_data_start_row, column=1, value="Compliance Level")
    ws.cell(row=pie_data_start_row, column=2, value="Count")
    ws.cell(row=pie_data_start_row, column=1).font = Font(bold=True)
    ws.cell(row=pie_data_start_row, column=2).font = Font(bold=True)

    level_counts = {0: 0, 1: 0, 2: 0, 3: 0}
    for his_item in manager.list_result:
        vg_value = his_item.get_matrix_value(MatrixItem.V_G)
        warn_level = manager.check_warning_level(MatrixItem.V_G, vg_value)
        level_counts[warn_level] = level_counts.get(warn_level, 0) + 1

    labels = {0: "Pass (Green)", 1: "Warning (Yellow)", 2: "Caution (Orange)", 3: "Fail (Red)"}
    pie_colors = {0: "00B050", 1: "FFFF00", 2: "FFA500", 3: "FF0000"}
    pie_row = pie_data_start_row + 1
    for level in range(4):
        ws.cell(row=pie_row, column=1, value=labels[level])
        ws.cell(row=pie_row, column=2, value=level_counts[level])
        pie_row += 1

    pie_data_end_row = pie_row - 1

    pie_chart = PieChart()
    pie_chart.title = "Overall Compliance Breakdown"
    pie_chart.style = 10
    pie_chart.width = 18
    pie_chart.height = 14

    pie_cats = Reference(ws, min_col=1, min_row=pie_data_start_row + 1, max_row=pie_data_end_row)
    pie_vals = Reference(ws, min_col=2, min_row=pie_data_start_row, max_row=pie_data_end_row)
    pie_chart.add_data(pie_vals, titles_from_data=True)
    pie_chart.set_categories(pie_cats)

    # 파이 슬라이스 색상 적용
    pie_series = pie_chart.series[0]
    for idx, level in enumerate(range(4)):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = pie_colors[level]
        pie_series.data_points.append(pt)

    ws.add_chart(pie_chart, "E" + str(pie_data_start_row))

    # 열 너비 조정
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
