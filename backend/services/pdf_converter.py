"""PDF 변환 유틸리티

DOCX → PDF 변환 및 간단한 PDF 생성 기능을 제공합니다.
- Windows: docx2pdf (Microsoft Word COM) 우선 사용
- Fallback: reportlab 기반 직접 PDF 생성
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path
from typing import Optional

# Auto-install dependencies
def _ensure_package(name: str, pip_name: str = None):
    try:
        __import__(name)
    except ImportError:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", pip_name or name, "-q"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )

_HAS_DOCX2PDF = False
_HAS_REPORTLAB = False

try:
    import docx2pdf
    _HAS_DOCX2PDF = True
except ImportError:
    pass

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _HAS_REPORTLAB = True
except ImportError:
    pass


def docx_to_pdf(docx_path: Path, pdf_path: Optional[Path] = None) -> Path:
    """Convert DOCX to PDF.

    Uses docx2pdf (Word COM) on Windows, auto-installs if missing.
    Returns the output PDF path.
    """
    docx_path = Path(docx_path).resolve()
    if not docx_path.exists():
        raise FileNotFoundError(f"DOCX not found: {docx_path}")

    if pdf_path is None:
        pdf_path = docx_path.with_suffix(".pdf")
    pdf_path = Path(pdf_path).resolve()
    pdf_path.parent.mkdir(parents=True, exist_ok=True)

    global _HAS_DOCX2PDF
    if not _HAS_DOCX2PDF:
        try:
            _ensure_package("docx2pdf")
            import docx2pdf as _d2p
            _HAS_DOCX2PDF = True
        except Exception:
            raise RuntimeError(
                "docx2pdf 설치 실패. Microsoft Word가 필요합니다. "
                "대안: xlsx_to_pdf() 또는 generate_report_pdf()를 사용하세요."
            )

    import docx2pdf as _d2p
    _d2p.convert(str(docx_path), str(pdf_path))
    return pdf_path


def xlsx_to_pdf(xlsx_path: Path, pdf_path: Optional[Path] = None, sheet_name: str = None) -> Path:
    """Convert Excel data to PDF using reportlab.

    Reads Excel with openpyxl and renders as PDF table.
    """
    _ensure_reportlab()

    xlsx_path = Path(xlsx_path).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel not found: {xlsx_path}")

    if pdf_path is None:
        pdf_path = xlsx_path.with_suffix(".pdf")
    pdf_path = Path(pdf_path).resolve()

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Extract data
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(cell or "") for cell in row])

    if not data:
        raise ValueError("Empty worksheet")

    # Build PDF
    styles = _get_styles()
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                           leftMargin=15*mm, rightMargin=15*mm,
                           topMargin=20*mm, bottomMargin=20*mm)

    elements = []
    elements.append(Paragraph(xlsx_path.stem, styles["title"]))
    elements.append(Spacer(1, 10*mm))

    # Split into page-sized chunks if too many columns
    max_cols = 8
    for col_start in range(0, len(data[0]) if data else 0, max_cols):
        col_end = min(col_start + max_cols, len(data[0]))
        chunk = [row[col_start:col_end] for row in data[:200]]  # limit rows

        # Wrap long text
        wrapped = []
        for row in chunk:
            wrapped.append([
                Paragraph(cell[:100], styles["cell"]) if len(cell) > 30
                else cell for cell in row
            ])

        table = Table(wrapped, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0052CC")),
            ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#FFFFFF"), HexColor("#F5F5F5")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)

        if col_end < len(data[0]):
            elements.append(PageBreak())

    doc.build(elements)
    return pdf_path


def generate_report_pdf(
    title: str,
    sections: list,
    pdf_path: Path,
    subtitle: str = "",
) -> Path:
    """Generate a structured PDF report from sections.

    Args:
        title: Report title
        sections: List of dicts with keys: heading, content (str or list of rows)
        pdf_path: Output path
        subtitle: Optional subtitle

    Returns: Path to generated PDF
    """
    _ensure_reportlab()

    pdf_path = Path(pdf_path).resolve()
    pdf_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _get_styles()
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                           leftMargin=20*mm, rightMargin=20*mm,
                           topMargin=25*mm, bottomMargin=20*mm)

    elements = []

    # Title page
    elements.append(Spacer(1, 40*mm))
    elements.append(Paragraph(title, styles["title"]))
    if subtitle:
        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph(subtitle, styles["subtitle"]))
    elements.append(Spacer(1, 10*mm))

    from datetime import datetime
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles["cell"]
    ))
    elements.append(PageBreak())

    # Sections
    for section in sections:
        heading = section.get("heading", "")
        content = section.get("content", "")

        if heading:
            elements.append(Paragraph(heading, styles["heading"]))
            elements.append(Spacer(1, 3*mm))

        if isinstance(content, str):
            for para in content.split("\n"):
                if para.strip():
                    elements.append(Paragraph(para.strip(), styles["body"]))
                    elements.append(Spacer(1, 2*mm))
        elif isinstance(content, list) and content:
            # Table data
            table = Table(content, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0052CC")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#FFFFFF"), HexColor("#F8F8F8")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 5*mm))

    doc.build(elements)
    return pdf_path


def _ensure_reportlab():
    global _HAS_REPORTLAB
    if not _HAS_REPORTLAB:
        _ensure_package("reportlab")
        # Re-import after install
        global A4, mm, SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        global getSampleStyleSheet, ParagraphStyle, HexColor, pdfmetrics, TTFont
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.colors import HexColor
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        _HAS_REPORTLAB = True


def _get_styles():
    """Get PDF styles with Korean font support."""
    styles = getSampleStyleSheet()

    # Try to register Korean font
    font_name = "Helvetica"
    for font_path in [
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    ]:
        try:
            if Path(font_path).exists():
                pdfmetrics.registerFont(TTFont("KoreanFont", font_path))
                font_name = "KoreanFont"
                break
        except Exception:
            continue

    custom = {
        "title": ParagraphStyle(
            "dr_title", parent=styles["Title"],
            fontName=font_name, fontSize=18, leading=24,
            textColor=HexColor("#172B4D"),
        ),
        "subtitle": ParagraphStyle(
            "dr_subtitle", parent=styles["Normal"],
            fontName=font_name, fontSize=12, leading=16,
            textColor=HexColor("#6B778C"),
        ),
        "heading": ParagraphStyle(
            "dr_heading", parent=styles["Heading2"],
            fontName=font_name, fontSize=13, leading=18,
            textColor=HexColor("#0052CC"), spaceBefore=12,
        ),
        "body": ParagraphStyle(
            "dr_body", parent=styles["Normal"],
            fontName=font_name, fontSize=9, leading=13,
        ),
        "cell": ParagraphStyle(
            "dr_cell", parent=styles["Normal"],
            fontName=font_name, fontSize=7, leading=10,
        ),
    }
    # Return custom styles as dict-like access
    return custom
